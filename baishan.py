import time
import urllib
import datetime
from re import split
from urllib.parse import urlencode
import requests, json
from openpyxl import load_workbook


def main1():
    api = "https://cdn.api.baishan.com/v2/domain/list?token=d54206623dfe730a1dddf12c1f4c0552&page_size=500"
    result = requests.get(api).json()
    result = json.dumps(result)
    load_data = json.loads(result)
    data = load_data.get("data").get("list")
    n = 0
    s = set()
    for i in data:
        result2 = (i.get("domain"))
        n = n + 1
        print("通过api获取第%d条域名:%s" % (n, result2))
        str = result2.split(".")
        newstr = str[-2] + "." + str[-1]
        print("处理后的域名为：%s" % newstr)
        s.add(newstr)
    print("通过api获取到的域名记录有%d条" % n)
    print("经过清除重复处理后的记录有%d条" % len(s))
    return s


# --------------计算里到期时间utils------------------
def time_passed(value):
    str_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timeArray = time.strptime(str_time, "%Y-%m-%d %H:%M:%S")
    timeStamp = int(time.mktime(timeArray))
    print("当前时间戳为：", timeStamp)
    str_time2 = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    timeArray2 = time.strptime(value, "%Y-%m-%d %H:%M:%S")
    timeStamp2 = int(time.mktime(timeArray2))
    print("到期时间戳为：", timeStamp2)
    # if timeStamp2>timeStamp:
    day_num = (timeStamp2 - timeStamp) / (24 * 60 * 60)
    return day_num


# --------------查询到期时间并调用utils计算到期时间--------------------
def due(hostname):
    url = 'http://api.k780.com'
    params = {
        'app': 'domain.whois',
        'domain': hostname,
        'appkey': '51781',
        'sign': '35c71b5d38215644068b602030ddde7b',
        'format': 'json',
    }
    params = urlencode(params)

    f = urllib.request.urlopen('%s?%s' % (url, params))
    nowapi_call = f.read()
    # print content
    a_result = json.loads(nowapi_call)
    if a_result:

        if a_result['success'] != '0':
            if a_result['result']['status'] == 'WAIT_PROCESS':
                print("%s:等待系统处理（预计两分钟后自动重试）")
                return "系统处理失败"
                # due(hostname)
                # time.sleep(130)
                # 处理四次，还不成功则直接返回

            if a_result['result']['status'] == 'ALREADY_WHOIS':
                t2 = a_result['result']['dom_expdate']
                print("获取到期时间成功" + hostname)
                return (time_passed(t2))

            if a_result['result']['status'] == 'NOT_REGISTER':
                print("未注册域名" + hostname)
                return "未注册域名"
            if a_result['result']['status'] == 'BE_RETAINED':
                print("域名被保留" + hostname)
                return "域名被保留"


        else:
            # 如果超额，70分钟后自动继续
            if a_result['msgid'] == '1000701':
                print("超额，程序将在70分钟后继续执行,请无需关闭........")
                time.sleep(70 * 60)
                due(hostname)
            if a_result['msgid'] == '10018':
                return "未识别域名"
            print(a_result['msgid'] + ' ' + a_result['msg'])
            print("查询失败，正在重试。。。。。。")
            due(hostname)
    else:
        print('Request nowapi fail.')
        print("查询失败，正在重试。。。。。。")
        due(hostname)
    # time.sleep(5)


# 把成功或者失败的域名分别写入excel
def write(hostname):
    num = 0
    from openpyxl import Workbook
    wb = load_workbook('h.xlsx')
    ws = wb.active
    ws['A1'] = "有效域名"
    ws['B1'] = "到期时间"
    ws['C1'] = "无效域名"
    ws['D1'] = "到期时间"
    i = 1
    d = due(hostname)
    try:
        d = int(d)
        if d > 0:
            while True:
                i = i + 1
                if (ws["A" + str(i)].value) == None:
                    ws["A" + str(i)] = hostname
                    ws["B" + str(i)] = str(d) + "天"
                    break
        else:
            while True:
                i = i + 1
                if (ws["C" + str(i)].value) == None:
                    ws["C" + str(i)] = hostname
                    ws["D" + str(i)] = str(d) + "天"
                    break
    except:
        while True:
            i = i + 1
            if (ws["C" + str(i)].value) == None:
                ws["C" + str(i)] = hostname
                ws["D" + str(i)] = d
                break

    wb.save("h.xlsx")


#
#
#
#
#
#
def fun1():
    s = main1();
    num = 0
    for i in s:
        num = num + 1
        print("写入第%d条数据,域名为%s" % (num, i))
        try:
            write(i)
            print("写入成功！！！")

        except Exception as e:
            print("写入失败！！")


# print("处理失败的域名为：%s,以下是错误信息" % i)
# print(e)

# 每二十四小时运行两次
#     time.sleep(12*60*60)
#     fun1()
fun1();

# due("sfaa.tv")
